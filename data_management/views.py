from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse 
from .models import UploadHistory
from setup.models import GLTransaction 
from .forms import IncomeStatementFilterForm 
import io 
from openpyxl import Workbook 
from django.utils.formats import number_format 
from django.shortcuts import render
# FIX: Import FundTransaction model
from setup.models import GLTransaction, FundTransaction
from .forms import IncomeStatementFilterForm, HistoricalDataUploadForm # ADDED HistoricalDataUploadForm



# Helper function to generate mock periods for headers
def generate_mock_periods(period_type):
    if period_type == 'monthly':
        return ['Jan', 'Feb', 'Mar', 'Apr']
    elif period_type == 'quarterly':
        return ['Q1', 'Q2', 'Q3', 'Q4']
    elif period_type == 'half_yearly':
        return ['H1', 'H2']
    else: # annual
        return ['Current Period', 'Previous Period']

@login_required
def historical_data_view(request):
    
    upload_history = UploadHistory.objects.filter(uploaded_by=request.user).order_by('-upload_date')[:10]
    recent_gl_transactions = GLTransaction.objects.all().order_by('-created_at')[:5]
    recent_fund_transactions = FundTransaction.objects.all().order_by('-created_at')[:5]
    
    if request.method == 'POST':
        upload_form = HistoricalDataUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            uploaded_file = upload_form.cleaned_data['excel_file']
            upload_type = upload_form.cleaned_data['upload_type']
            
            # --- Simulation of Import Logic ---
            # In a real app, this is where you'd call a parser function
            # that routes data based on upload_type (e.g., to GLTransaction, FundTransaction tables)
            
            # Simulate logging success
            UploadHistory.objects.create(
                file_name=uploaded_file.name,
                uploaded_by=request.user,
                record_count=100, # Mock count
                status=f'Success: Routed to {upload_type} table'
            )
            # End Simulation ---

            # Redirect after POST to prevent resubmission
            return redirect('data_management:historical_data') 
    else:
        upload_form = HistoricalDataUploadForm()


    context = {
        'upload_history': upload_history,
        'recent_gl_transactions': recent_gl_transactions, 
        'recent_fund_transactions': recent_fund_transactions, 
        'upload_form': upload_form, # Pass the form to the template
    }
    return render(request, 'data_management/historical_data.html', context)


@login_required
def download_excel_template(request, template_type):
    """Generates and serves an Excel file template based on the type requested."""
    
    # Define column headers for different templates
    templates = {
        'gl_transactions': {
            'filename': 'GL_Transactions_Template.xlsx',
            'headers': [
                'transaction_date (YYYY-MM-DD)', 'gl_account_code', 'description', 
                'journal_type', 'document_no', 'reference_no', 'entity_code', 
                'cost_center_code', 'project_code', 'currency_code', 
                'exchange_rate', 'debit', 'credit'
            ]
        },
        'gl_accounts': {
            'filename': 'GL_Chart_of_Accounts_Template.xlsx',
            'headers': [
                'gl_account_code', 'gl_account_name', 'category', 'sub_category',
                'financial_statement (Income Statement/Balance Sheet/Cash Flow)', 
                'account_type (Account/Header)', 'is_postable (TRUE/FALSE)', 
                'parent_account_code', 'normal_balance (Credit/Debit)', 'active_flag (TRUE/FALSE)'
            ]
        },
        'date_table': {
            'filename': 'Date_Detail_Template.xlsx',
            'headers': ['date (YYYY-MM-DD)']
        },
        # NEW: RSA Fund Template
        'rsa_fund': {
            'filename': 'RSA_Fund_Historical_Template.xlsx',
            'headers': [
                'transaction_date (YYYY-MM-DD)', 'rsa_fund_name', 'entity_code', 
                'contributions', 'withdrawals', 'balance'
            ]
        },
        # NEW: Managed Fund Template
        'managed_fund': {
            'filename': 'Managed_Fund_Historical_Template.xlsx',
            'headers': [
                'transaction_date (YYYY-MM-DD)', 'managed_fund_name', 'entity_code', 
                'investment_value', 'contributions', 'withdrawals'
            ]
        },
    }

    template_info = templates.get(template_type)

    if not template_info:
        return HttpResponse("Invalid template type requested.", status=404)

    # 1. Create workbook and add headers
    wb = Workbook()
    ws = wb.active
    ws.title = template_info['filename'].replace('.xlsx', '')
    ws.append(template_info['headers'])

    # 2. Prepare in-memory file (BytesIO is essential for Django file serving)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 3. Create HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={template_info["filename"]}'
    
    return response

@login_required
def income_statement_view(request):
    """Renders the Income Statement page with performance cards and financial data, respecting filters."""
    
    # Initialize form with GET data if filters are applied, otherwise use unbound form
    filter_form = IncomeStatementFilterForm(request.GET)
    
    # Default Period Settings
    period_title = 'FY 2024'
    previous_period_title = 'FY 2023'
    applied_filters = {}
    period_labels = ['Current Period', 'Previous Period']
    
    # --- Simulate Filter Application ---
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
        
        # Simulate change in report title based on date filter
        if start_date and end_date:
            period_title = f"{start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
            previous_period_title = "Prior Period"
        elif start_date:
            period_title = f"Since {start_date.strftime('%b %d, %Y')}"
            previous_period_title = "Prior Period"
            
        # Collect only non-empty filters for display
        applied_filters = {
            filter_form.fields[k].label: v 
            for k, v in filter_form.cleaned_data.items() 
            if v not in (None, '', False) # Filter out None, empty string, and False
        }

    # Mock Performance Data and Financial Data (Remains static for now)
    performance_cards = [
        {'title': 'Total Revenue', 'value': '₦1.24B', 'change': '+15.2%', 'trend': 'up', 'icon': 'fas fa-arrow-up', 'color': 'success'},
        {'title': 'Operating Expenses', 'value': '₦510M', 'change': '-2.5%', 'trend': 'down', 'icon': 'fas fa-arrow-down', 'color': 'danger'},
        {'title': 'Net Profit', 'value': '₦320M', 'change': '+8.9%', 'trend': 'up', 'icon': 'fas fa-chart-line', 'color': 'primary'},
        {'title': 'Profit Margin', 'value': '25.8%', 'change': '+1.1%', 'trend': 'up', 'icon': 'fas fa-percentage', 'color': 'info'},
    ]
    
    # Mock Data adapted for standard display
    financial_data = [
        {'description': 'REVENUE', 'current': 1240000000, 'previous': 1076000000, 'type': 'header', 'periods': {'Current Period': 1240000000, 'Previous Period': 1076000000}},
        {'description': 'Management Fees Income', 'current': 950000000, 'previous': 820000000, 'type': 'account', 'periods': {'Current Period': 950000000, 'Previous Period': 820000000}},
        {'description': 'Administrative Fees Income', 'current': 290000000, 'previous': 256000000, 'type': 'account', 'periods': {'Current Period': 290000000, 'Previous Period': 256000000}},
        {'description': 'TOTAL REVENUE', 'current': 1240000000, 'previous': 1076000000, 'type': 'subtotal', 'periods': {'Current Period': 1240000000, 'Previous Period': 1076000000}},
        
        {'description': 'OPERATING EXPENSES', 'current': 510000000, 'previous': 523000000, 'type': 'header', 'periods': {'Current Period': 510000000, 'Previous Period': 523000000}},
        {'description': 'Staff Costs', 'current': 300000000, 'previous': 310000000, 'type': 'account', 'periods': {'Current Period': 300000000, 'Previous Period': 310000000}},
        {'description': 'Depreciation', 'current': 55000000, 'previous': 52000000, 'type': 'account', 'periods': {'Current Period': 55000000, 'Previous Period': 52000000}},
        {'description': 'Administrative Expenses', 'current': 155000000, 'previous': 161000000, 'type': 'account', 'periods': {'Current Period': 155000000, 'Previous Period': 161000000}},
        {'description': 'TOTAL OPERATING EXPENSES', 'current': 510000000, 'previous': 523000000, 'type': 'subtotal', 'periods': {'Current Period': 510000000, 'Previous Period': 523000000}},
        
        {'description': 'FINANCE INCOME', 'current': 120000000, 'previous': 100000000, 'type': 'header', 'periods': {'Current Period': 120000000, 'Previous Period': 100000000}},
        {'description': 'Interest Income', 'current': 120000000, 'previous': 100000000, 'type': 'account', 'periods': {'Current Period': 120000000, 'Previous Period': 100000000}},
        
        {'description': 'NET PROFIT BEFORE TAX', 'current': 850000000, 'previous': 653000000, 'type': 'major_total', 'periods': {'Current Period': 850000000, 'Previous Period': 653000000}},
    ]
    
    context = {
        'filter_form': filter_form,
        'applied_filters': applied_filters,
        'performance_cards': performance_cards,
        'financial_data': financial_data,
        'period_labels': period_labels, # Pass dynamic labels for I.S.
        'period': period_title,
        'previous_period': previous_period_title,
        'report_type': 'Income Statement',
        'period_prefix': 'For the Period Ended:',
    }
    return render(request, 'data_management/income_statement.html', context)


@login_required
def export_income_statement_excel(request):
    """Exports the mock Income Statement data to an Excel file."""
    # ... (remains unchanged) ...
    # This export should be updated similarly but kept simple for now since I.S. wasn't the target of the multi-period change.
    financial_data = [
        {'description': 'REVENUE', 'current': 1240000000, 'previous': 1076000000, 'type': 'header'},
        {'description': 'Management Fees Income', 'current': 950000000, 'previous': 820000000, 'type': 'account'},
        {'description': 'Administrative Fees Income', 'current': 290000000, 'previous': 256000000, 'type': 'account'},
        {'description': 'TOTAL REVENUE', 'current': 1240000000, 'previous': 1076000000, 'type': 'subtotal'},
        
        {'description': 'OPERATING EXPENSES', 'current': 510000000, 'previous': 523000000, 'type': 'header'},
        {'description': 'Staff Costs', 'current': 300000000, 'previous': 310000000, 'type': 'account'},
        {'description': 'Depreciation', 'current': 55000000, 'previous': 52000000, 'type': 'account'},
        {'description': 'Administrative Expenses', 'current': 155000000, 'previous': 161000000, 'type': 'account'},
        {'description': 'TOTAL OPERATING EXPENSES', 'current': 510000000, 'previous': 523000000, 'type': 'subtotal'},
        
        {'description': 'FINANCE INCOME', 'current': 120000000, 'previous': 100000000, 'type': 'header'},
        {'description': 'Interest Income', 'current': 120000000, 'previous': 100000000, 'type': 'account'},
        
        {'description': 'NET PROFIT BEFORE TAX', 'current': 850000000, 'previous': 653000000, 'type': 'major_total'},
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement FY2024"
    
    # Headers
    ws.append(['Description', 'FY 2024 (Current)', 'FY 2023 (Previous)'])
    
    # Data Rows
    for item in financial_data:
        # Use number_format to ensure Django locale settings (if any) are respected
        current_amount = item['current'] if item['current'] is not None else ''
        previous_amount = item['previous'] if item['previous'] is not None else ''
        
        ws.append([
            item['description'],
            current_amount,
            previous_amount
        ])

    # Prepare in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Income_Statement_FY2024.xlsx'
    
    return response


@login_required
def balance_sheet_view(request):
    """Renders the Balance Sheet page with performance cards and financial data, respecting filters."""
    
    filter_form = IncomeStatementFilterForm(request.GET)
    
    # Balance Sheet uses an 'As At' period, often the end date
    period_title = 'December 31, 2024'
    previous_period_title = 'December 31, 2023'
    applied_filters = {}
    period_labels = ['Current Period', 'Previous Period']
    
    # --- Simulate Filter Application (Only date changes report header) ---
    if filter_form.is_valid():
        end_date = filter_form.cleaned_data.get('end_date')
        if end_date:
            period_title = end_date.strftime('%B %d, %Y')
            previous_period_title = "Prior Year"
        
        applied_filters = {
            filter_form.fields[k].label: v 
            for k, v in filter_form.cleaned_data.items() 
            if v not in (None, '', False)
        }

    # Mock Performance Data
    performance_cards = [
        {'title': 'Total Assets', 'value': '₦2.5B', 'change': '+5.0%', 'trend': 'up', 'icon': 'fas fa-arrow-up', 'color': 'success'},
        {'title': 'Total Liabilities', 'value': '₦1.1B', 'change': '-1.2%', 'trend': 'down', 'icon': 'fas fa-arrow-down', 'color': 'danger'},
        {'title': 'Total Equity', 'value': '₦1.4B', 'change': '+10.5%', 'trend': 'up', 'icon': 'fas fa-chart-line', 'color': 'primary'},
        {'title': 'Current Ratio', 'value': '2.1x', 'change': '-0.1x', 'trend': 'down', 'icon': 'fas fa-percentage', 'color': 'warning'},
    ]
    
    # Mock Balance Sheet Data (Assets = Liabilities + Equity)
    financial_data = [
        {'description': 'ASSETS', 'current': 2500000000, 'previous': 2380000000, 'type': 'section_header', 'periods': {'Current Period': 2500000000, 'Previous Period': 2380000000}},
        {'description': 'Non-Current Assets', 'current': 1500000000, 'previous': 1400000000, 'type': 'header', 'periods': {'Current Period': 1500000000, 'Previous Period': 1400000000}},
        {'description': 'Property, Plant & Equipment', 'current': 1450000000, 'previous': 1350000000, 'type': 'account', 'periods': {'Current Period': 1450000000, 'Previous Period': 1350000000}},
        {'description': 'Current Assets', 'current': 1000000000, 'previous': 980000000, 'type': 'header', 'periods': {'Current Period': 1000000000, 'Previous Period': 980000000}},
        {'description': 'Cash & Bank Balances', 'current': 450000000, 'previous': 420000000, 'type': 'account', 'periods': {'Current Period': 450000000, 'Previous Period': 420000000}},
        {'description': 'Trade Receivables', 'current': 550000000, 'previous': 560000000, 'type': 'account', 'periods': {'Current Period': 550000000, 'Previous Period': 560000000}},
        {'description': 'TOTAL ASSETS', 'current': 2500000000, 'previous': 2380000000, 'type': 'major_total', 'periods': {'Current Period': 2500000000, 'Previous Period': 2380000000}},

        {'description': 'LIABILITIES', 'current': 1100000000, 'previous': 1113000000, 'type': 'section_header', 'periods': {'Current Period': 1100000000, 'Previous Period': 1113000000}},
        {'description': 'Current Liabilities', 'current': 550000000, 'previous': 560000000, 'type': 'header', 'periods': {'Current Period': 550000000, 'Previous Period': 560000000}},
        {'description': 'Trade Payables', 'current': 300000000, 'previous': 310000000, 'type': 'account', 'periods': {'Current Period': 300000000, 'Previous Period': 310000000}},
        {'description': 'Accrued Liabilities', 'current': 250000000, 'previous': 250000000, 'type': 'account', 'periods': {'Current Period': 250000000, 'Previous Period': 250000000}},
        {'description': 'Non-Current Liabilities', 'current': 550000000, 'previous': 553000000, 'type': 'header', 'periods': {'Current Period': 550000000, 'Previous Period': 553000000}},
        {'description': 'Long-Term Borrowings', 'current': 550000000, 'previous': 553000000, 'type': 'account', 'periods': {'Current Period': 550000000, 'Previous Period': 553000000}},
        {'description': 'TOTAL LIABILITIES', 'current': 1100000000, 'previous': 1113000000, 'type': 'subtotal', 'periods': {'Current Period': 1100000000, 'Previous Period': 1113000000}},

        {'description': 'EQUITY', 'current': 1400000000, 'previous': 1267000000, 'type': 'section_header', 'periods': {'Current Period': 1400000000, 'Previous Period': 1267000000}},
        {'description': 'Share Capital', 'current': 800000000, 'previous': 800000000, 'type': 'account', 'periods': {'Current Period': 800000000, 'Previous Period': 800000000}},
        {'description': 'Retained Earnings', 'current': 600000000, 'previous': 467000000, 'type': 'account', 'periods': {'Current Period': 600000000, 'Previous Period': 467000000}},
        {'description': 'TOTAL EQUITY', 'current': 1400000000, 'previous': 1267000000, 'type': 'subtotal', 'periods': {'Current Period': 1400000000, 'Previous Period': 1267000000}},
        
        {'description': 'TOTAL LIABILITIES & EQUITY', 'current': 2500000000, 'previous': 2380000000, 'type': 'major_total', 'periods': {'Current Period': 2500000000, 'Previous Period': 2380000000}},
    ]
    
    context = {
        'filter_form': filter_form,
        'applied_filters': applied_filters,
        'performance_cards': performance_cards,
        'financial_data': financial_data,
        'period_labels': period_labels, # Pass dynamic labels for B.S.
        'period': period_title,
        'report_type': 'Balance Sheet',
        'period_prefix': 'As At:',
    }
    return render(request, 'data_management/balance_sheet.html', context)


@login_required
def export_balance_sheet_excel(request):
    # ... (remains unchanged) ...
    # ... (omitted for brevity) ...

    # Prepare in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Balance_Sheet_FY2024.xlsx'
    
    return response


# --- NEW Cash Flow Views ---

@login_required
def cash_flow_view(request):
    """Renders the Cash Flow Statement page with performance cards and financial data, respecting filters."""
    
    filter_form = IncomeStatementFilterForm(request.GET)
    
    period_title = 'FY 2024'
    period_prefix = 'For the Period Ended:'
    applied_filters = {}
    reporting_period = filter_form.data.get('reporting_period', 'annual') # Get the chosen period
    
    # --- Simulate Filter Application ---
    if filter_form.is_valid():
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
        reporting_period = filter_form.cleaned_data.get('reporting_period') or 'annual'
        
        if start_date and end_date and reporting_period == 'annual':
            period_title = f"{start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
            
        applied_filters = {
            filter_form.fields[k].label: v 
            for k, v in filter_form.cleaned_data.items() 
            if v not in (None, '', False) and k != 'reporting_period'
        }
        
        if reporting_period != 'annual':
             applied_filters['Period Type'] = dict(filter_form.fields['reporting_period'].choices).get(reporting_period)


    # Mock Performance Data
    performance_cards = [
        {'title': 'Net Operating Cash', 'value': '₦730M', 'change': '+15%', 'trend': 'up', 'icon': 'fas fa-briefcase', 'color': 'success'},
        {'title': 'Cash from Investing', 'value': '-₦120M', 'change': '+5%', 'trend': 'up', 'icon': 'fas fa-chart-line', 'color': 'primary'},
        {'title': 'Net Change in Cash', 'value': '₦680M', 'change': '+20%', 'trend': 'up', 'icon': 'fas fa-balance-scale', 'color': 'info'},
        {'title': 'Ending Cash Balance', 'value': '₦1.1B', 'change': '+18%', 'trend': 'up', 'icon': 'fas fa-university', 'color': 'warning'},
    ]
    
    # Generate headers and adapt data for time series
    period_labels = generate_mock_periods(reporting_period)
    
    # Mock Data adapted for multi-period display
    data_template = [
        {'description': 'CASH FLOW FROM OPERATING ACTIVITIES', 'data': [730000000, 635000000], 'type': 'section_header'},
        {'description': 'Cash received from customers', 'data': [1400000000, 1250000000], 'type': 'account'},
        {'description': 'Cash paid to suppliers and employees', 'data': [-670000000, -615000000], 'type': 'account'},
        {'description': 'NET CASH FROM OPERATING ACTIVITIES', 'data': [730000000, 635000000], 'type': 'subtotal'},
        
        {'description': 'CASH FLOW FROM INVESTING ACTIVITIES', 'data': [-120000000, -114000000], 'type': 'section_header'},
        {'description': 'Purchase of Property, Plant & Equipment', 'data': [-150000000, -140000000], 'type': 'account'},
        {'description': 'NET CASH USED IN INVESTING ACTIVITIES', 'data': [-120000000, -114000000], 'type': 'subtotal'},
        
        {'description': 'NET INCREASE IN CASH AND CASH EQUIVALENTS', 'data': [680000000, 576000000], 'type': 'major_total'},
        {'description': 'CASH AND CASH EQUIVALENTS AT END OF PERIOD', 'data': [1100000000, 956000000], 'type': 'major_total'},
    ]
    
    financial_data = []

    for item in data_template:
        new_item = {'description': item['description'], 'type': item['type'], 'periods': {}}
        
        if reporting_period != 'annual':
            # Simulate split for monthly/quarterly/half-yearly
            num_periods = len(period_labels)
            current_base = item['data'][0] / num_periods if num_periods > 0 else 0
            
            for i, label in enumerate(period_labels):
                # Mock small variance across periods
                new_item['periods'][label] = current_base + (i * 1000000) 
        else:
            # Standard annual structure
            new_item['periods']['Current Period'] = item['data'][0]
            new_item['periods']['Previous Period'] = item['data'][1]
            
        financial_data.append(new_item)


    context = {
        'filter_form': filter_form,
        'applied_filters': applied_filters,
        'performance_cards': performance_cards,
        'financial_data': financial_data,
        'period_labels': period_labels, # Pass dynamic labels
        'period': period_title,
        'report_type': 'Cash Flow Statement',
        'period_prefix': period_prefix,
    }
    return render(request, 'data_management/cash_flow.html', context)


@login_required
def export_cash_flow_excel(request):
    """Exports the mock Cash Flow data to an Excel file."""
    
    # Retrieve reporting period filter (default to annual)
    reporting_period = request.GET.get('reporting_period', 'annual')
    period_labels = generate_mock_periods(reporting_period)
    
    # Mock Data (Direct Method Simulation)
    financial_data_raw = [
        # Same structure as data_template in cash_flow_view
        {'description': 'CASH FLOW FROM OPERATING ACTIVITIES', 'data': [730000000, 635000000], 'type': 'section_header'},
        {'description': 'Cash received from customers', 'data': [1400000000, 1250000000], 'type': 'account'},
        {'description': 'Cash paid to suppliers and employees', 'data': [-670000000, -615000000], 'type': 'account'},
        {'description': 'NET CASH FROM OPERATING ACTIVITIES', 'data': [730000000, 635000000], 'type': 'subtotal'},
        
        {'description': 'CASH FLOW FROM INVESTING ACTIVITIES', 'data': [-120000000, -114000000], 'type': 'section_header'},
        {'description': 'Purchase of Property, Plant & Equipment', 'data': [-150000000, -140000000], 'type': 'account'},
        {'description': 'NET CASH USED IN INVESTING ACTIVITIES', 'data': [-120000000, -114000000], 'type': 'subtotal'},
        
        {'description': 'NET INCREASE IN CASH AND CASH EQUIVALENTS', 'data': [680000000, 576000000], 'type': 'major_total'},
        {'description': 'CASH AND CASH EQUIVALENTS AT END OF PERIOD', 'data': [1100000000, 956000000], 'type': 'major_total'},
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement FY2024"
    
    # Dynamic Headers
    headers = ['Description'] + period_labels
    ws.append(headers)

    # Populate Data Rows (using the same mock logic as the view)
    for item in financial_data_raw:
        row = [item['description']]
        
        if reporting_period != 'annual':
            num_periods = len(period_labels)
            current_base = item['data'][0] / num_periods if num_periods > 0 else 0
            
            for i, label in enumerate(period_labels):
                row.append(current_base + (i * 1000000))
        else:
            row.append(item['data'][0])
            row.append(item['data'][1])
            
        ws.append(row)


    # Prepare in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Cash_Flow_Statement_{reporting_period.capitalize()}.xlsx'
    
    return response

# Mock data structure to simulate CSV content
MOCK_MANAGED_FUND_DATA = {
    'GUINNESS': {
        'AUM Closing Balance': 913523.56,
        'Contribution': 26122.28,
        'Payout': 9509.10,
        'Returns': 33838.47,
    },
    'CBN RETIREE': {
        'AUM Closing Balance': 26362188.82,
        'Contribution': 0,
        'Payout': 577931.51,
        'Returns': 1153444.67,
    },
    'NNPC': {
        'AUM Closing Balance': 51308240.58,
        'Contribution': 208245.94,
        'Payout': 0,
        'Returns': 2027588.83,
    }
}
MOCK_RSA_FUND_DATA = {
    'RSA Fund 1': {
        'AUM Closing Balance': 9553729.43,
        'Total PINs': 1149,
        'Active PINs': 837,
        'Avg Contribution (New)': 1162.34,
    },
    'RSA Fund 2': {
        'AUM Closing Balance': 571982124.79,
        'Total PINs': 520147,
        'Active PINs': 187929,
        'Avg Contribution (New)': 8.92,
    }
}

@login_required
def managed_fund_view(request):
    filter_form = IncomeStatementFilterForm(request.GET)
    
    # Mock periods based on CSV (Q4 2024 is the latest actual data point)
    period_labels = ['Q4 2023', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024']
    
    # Mock Performance Cards
    performance_cards = [
        {'title': 'Total AUM (Latest)', 'value': '₦83.9M', 'change': '+5%', 'trend': 'up', 'icon': 'fas fa-chart-pie', 'color': 'success'},
        {'title': 'Total Contributions', 'value': '₦250k', 'change': '+12%', 'trend': 'up', 'icon': 'fas fa-plus', 'color': 'primary'},
        {'title': 'Avg Quarterly Return', 'value': '1.5%', 'change': '-0.1%', 'trend': 'down', 'icon': 'fas fa-percentage', 'color': 'warning'},
    ]
    
    # Adapt mock data for table display
    financial_data = []
    
    for fund_name, metrics in MOCK_MANAGED_FUND_DATA.items():
        # AUM Closing Balance
        financial_data.append({
            'description': f"AUM Closing Balance - {fund_name}",
            'periods': {period: metrics['AUM Closing Balance'] + (i * 1000) for i, period in enumerate(period_labels)},
            'type': 'account',
            'is_major': False
        })
        # Contribution
        financial_data.append({
            'description': f"Contribution - {fund_name}",
            'periods': {period: metrics['Contribution'] + (i * 100) for i, period in enumerate(period_labels)},
            'type': 'account',
            'is_major': False
        })

    context = {
        'filter_form': filter_form,
        'performance_cards': performance_cards,
        'financial_data': financial_data,
        'period_labels': period_labels,
        'report_type': 'Managed Fund Historical Data',
        'period_prefix': 'Quarterly Breakdown:',
    }
    return render(request, 'data_management/managed_fund_report.html', context)


@login_required
def rsa_fund_view(request):
    filter_form = IncomeStatementFilterForm(request.GET)
    
    # Mock periods based on CSV (Q4 2024 is the latest actual data point)
    period_labels = ['Q4 2023', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024']
    
    # Mock Performance Cards
    performance_cards = [
        {'title': 'Total RSA AUM', 'value': '₦600B', 'change': '+12%', 'trend': 'up', 'icon': 'fas fa-shield-alt', 'color': 'info'},
        {'title': 'Total PIN Count', 'value': '645K', 'change': '+3%', 'trend': 'up', 'icon': 'fas fa-users', 'color': 'primary'},
        {'title': 'Avg Contrib. (New)', 'value': '₦550', 'change': '-10%', 'trend': 'down', 'icon': 'fas fa-arrow-down', 'color': 'danger'},
    ]
    
    # Adapt mock data for table display
    financial_data = []
    
    for fund_name, metrics in MOCK_RSA_FUND_DATA.items():
        # AUM
        financial_data.append({
            'description': f"AUM Closing Balance - {fund_name}",
            'periods': {period: metrics['AUM Closing Balance'] + (i * 500000) for i, period in enumerate(period_labels)},
            'type': 'header',
            'is_major': False
        })
        # Total PINs
        financial_data.append({
            'description': f"Total PINs - {fund_name}",
            'periods': {period: metrics['Total PINs'] + (i * 100) for i, period in enumerate(period_labels)},
            'type': 'account',
            'is_major': False
        })
        # Active PINs
        financial_data.append({
            'description': f"Active PINs - {fund_name}",
            'periods': {period: metrics['Active PINs'] + (i * 50) for i, period in enumerate(period_labels)},
            'type': 'account',
            'is_major': False
        })
        # Avg Contribution
        financial_data.append({
            'description': f"Avg Contribution (New) - {fund_name}",
            'periods': {period: metrics['Avg Contribution (New)'] + (i * 5) for i, period in enumerate(period_labels)},
            'type': 'account',
            'is_major': False
        })

    context = {
        'filter_form': filter_form,
        'performance_cards': performance_cards,
        'financial_data': financial_data,
        'period_labels': period_labels,
        'report_type': 'RSA Fund Historical Data',
        'period_prefix': 'Quarterly Breakdown:',
    }
    return render(request, 'data_management/rsa_fund_report.html', context)